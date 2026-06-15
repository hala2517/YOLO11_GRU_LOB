import datetime as dt
import io
from decimal import Decimal
from typing import Annotated, Any

from app.api.deps import get_current_user, get_db_session
from app.models.standard_process_chart import (
    StandardProcessChart,
    StandardProcessChartAttachment,
    StandardProcessChartLayout,
    StandardProcessChartStep,
)
from app.models.user import User
from app.schemas.standard_process_charts import (
    StandardProcessChartAttachmentCreate,
    StandardProcessChartCreate,
    StandardProcessChartLayoutUpsert,
    StandardProcessChartListItem,
    StandardProcessChartListResponse,
    StandardProcessChartResponse,
    StandardProcessChartStepCreate,
    StandardProcessChartStepResponse,
    StandardProcessChartStepUpdate,
    StandardProcessChartUpdate,
)
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

router = APIRouter(prefix="/standard-process-charts", tags=["standard-process-charts"])

DbSession = Annotated[AsyncSession, Depends(get_db_session)]
CurrentUser = Annotated[User, Depends(get_current_user)]

MAX_RAW_ROWS = 50
MAX_RAW_COLUMNS = 26
EMPTY_ROW_SCAN_LIMIT = 50


def _get_chart_or_404(chart: StandardProcessChart | None) -> StandardProcessChart:
    if chart is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="표준공정도를 찾을 수 없습니다.",
        )
    return chart


def _get_step_or_404(step: StandardProcessChartStep | None) -> StandardProcessChartStep:
    if step is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="표준공정도 공정을 찾을 수 없습니다.",
        )
    return step


async def _load_chart(session: AsyncSession, chart_id: int) -> StandardProcessChart:
    result = await session.execute(
        select(StandardProcessChart)
        .options(
            selectinload(StandardProcessChart.steps),
            selectinload(StandardProcessChart.layouts),
            selectinload(StandardProcessChart.attachments),
        )
        .where(
            StandardProcessChart.id == chart_id,
            StandardProcessChart.is_deleted.is_(False),
        )
    )
    return _get_chart_or_404(result.scalar_one_or_none())


def _chart_list_item(chart: StandardProcessChart) -> StandardProcessChartListItem:
    return StandardProcessChartListItem(
        id=chart.id,
        line_code=chart.line_code,
        product_code=chart.product_code,
        product_name=chart.product_name,
        name=chart.name,
        version=chart.version,
        status=chart.status,
        owner_department=chart.owner_department,
        effective_from=chart.effective_from,
        effective_to=chart.effective_to,
        note=chart.note,
        created_by=chart.created_by,
        is_deleted=chart.is_deleted,
        created_at=chart.created_at,
        updated_at=chart.updated_at,
        step_count=len(chart.steps),
    )


@router.post("", response_model=StandardProcessChartResponse, status_code=status.HTTP_201_CREATED)
async def create_standard_process_chart(
    body: StandardProcessChartCreate,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChart:
    data = body.model_dump(exclude={"steps", "layout", "attachments"})
    data["created_by"] = current_user.id
    chart = StandardProcessChart(**data)
    chart.steps = [StandardProcessChartStep(**step.model_dump()) for step in body.steps]
    if body.layout is not None:
        chart.layouts = [StandardProcessChartLayout(**body.layout.model_dump())]
    chart.attachments = [
        StandardProcessChartAttachment(**attachment.model_dump())
        for attachment in body.attachments
    ]

    session.add(chart)
    await session.commit()
    return await _load_chart(session, chart.id)


@router.get("", response_model=StandardProcessChartListResponse)
async def list_standard_process_charts(
    session: DbSession,
    current_user: CurrentUser,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    line_code: str | None = Query(default=None),
    product_code: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    include_archived: bool = Query(default=False),
) -> StandardProcessChartListResponse:
    conditions = [StandardProcessChart.is_deleted.is_(False)]
    if line_code is not None:
        conditions.append(StandardProcessChart.line_code == line_code)
    if product_code is not None:
        conditions.append(StandardProcessChart.product_code == product_code)
    if status_filter is not None:
        conditions.append(StandardProcessChart.status == status_filter)
    elif not include_archived:
        conditions.append(StandardProcessChart.status != "archived")

    total = (
        await session.execute(
            select(func.count()).select_from(StandardProcessChart).where(*conditions)
        )
    ).scalar_one()

    offset = (page - 1) * size
    charts = (
        await session.execute(
            select(StandardProcessChart)
            .options(selectinload(StandardProcessChart.steps))
            .where(*conditions)
            .order_by(
                StandardProcessChart.line_code.asc(),
                StandardProcessChart.product_code.asc(),
                StandardProcessChart.updated_at.desc(),
            )
            .offset(offset)
            .limit(size)
        )
    ).scalars().all()

    return StandardProcessChartListResponse(
        items=[_chart_list_item(chart) for chart in charts],
        total=total,
        page=page,
        size=size,
    )


@router.get("/lines/{line_code}", response_model=StandardProcessChartListResponse)
async def list_standard_process_charts_by_line(
    line_code: str,
    session: DbSession,
    current_user: CurrentUser,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    product_code: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
) -> StandardProcessChartListResponse:
    return await list_standard_process_charts(
        session=session,
        current_user=current_user,
        page=page,
        size=size,
        line_code=line_code,
        product_code=product_code,
        status_filter=status_filter,
        include_archived=False,
    )


@router.get("/{chart_id}", response_model=StandardProcessChartResponse)
async def get_standard_process_chart(
    chart_id: int,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChart:
    return await _load_chart(session, chart_id)


@router.put("/{chart_id}", response_model=StandardProcessChartResponse)
async def update_standard_process_chart(
    chart_id: int,
    body: StandardProcessChartUpdate,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChart:
    chart = await _load_chart(session, chart_id)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(chart, field, value)
    await session.commit()
    return await _load_chart(session, chart_id)


@router.delete("/{chart_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_standard_process_chart(
    chart_id: int,
    session: DbSession,
    current_user: CurrentUser,
) -> None:
    chart = await _load_chart(session, chart_id)
    chart.is_deleted = True
    await session.commit()


@router.post(
    "/{chart_id}/steps",
    response_model=StandardProcessChartStepResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_standard_process_chart_step(
    chart_id: int,
    body: StandardProcessChartStepCreate,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChartStep:
    await _load_chart(session, chart_id)
    step = StandardProcessChartStep(chart_id=chart_id, **body.model_dump())
    session.add(step)
    await session.commit()
    await session.refresh(step)
    return step


@router.put("/{chart_id}/steps/{step_id}", response_model=StandardProcessChartStepResponse)
async def update_standard_process_chart_step(
    chart_id: int,
    step_id: int,
    body: StandardProcessChartStepUpdate,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChartStep:
    result = await session.execute(
        select(StandardProcessChartStep).where(
            StandardProcessChartStep.id == step_id,
            StandardProcessChartStep.chart_id == chart_id,
        )
    )
    step = _get_step_or_404(result.scalar_one_or_none())
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(step, field, value)
    await session.commit()
    await session.refresh(step)
    return step


@router.delete("/{chart_id}/steps/{step_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_standard_process_chart_step(
    chart_id: int,
    step_id: int,
    session: DbSession,
    current_user: CurrentUser,
) -> None:
    result = await session.execute(
        select(StandardProcessChartStep).where(
            StandardProcessChartStep.id == step_id,
            StandardProcessChartStep.chart_id == chart_id,
        )
    )
    step = _get_step_or_404(result.scalar_one_or_none())
    await session.delete(step)
    await session.commit()


@router.put("/{chart_id}/layout", response_model=StandardProcessChartResponse)
async def upsert_standard_process_chart_layout(
    chart_id: int,
    body: StandardProcessChartLayoutUpsert,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChart:
    await _load_chart(session, chart_id)
    result = await session.execute(
        select(StandardProcessChartLayout).where(
            StandardProcessChartLayout.chart_id == chart_id,
            StandardProcessChartLayout.layout_type == body.layout_type,
        )
    )
    layout = result.scalar_one_or_none()
    if layout is None:
        layout = StandardProcessChartLayout(chart_id=chart_id, **body.model_dump())
        session.add(layout)
    else:
        for field, value in body.model_dump().items():
            setattr(layout, field, value)
    await session.commit()
    return await _load_chart(session, chart_id)


@router.post(
    "/{chart_id}/attachments",
    response_model=StandardProcessChartResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_standard_process_chart_attachment(
    chart_id: int,
    body: StandardProcessChartAttachmentCreate,
    session: DbSession,
    current_user: CurrentUser,
) -> StandardProcessChart:
    await _load_chart(session, chart_id)
    attachment = StandardProcessChartAttachment(chart_id=chart_id, **body.model_dump())
    session.add(attachment)
    await session.commit()
    return await _load_chart(session, chart_id)


def _json_value(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _cell_payload(cell: Cell) -> dict[str, Any]:
    return {
        "value": _json_value(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "is_date": cell.is_date,
        "coordinate": cell.coordinate,
    }


def _extract_raw_cells(sheet: Worksheet) -> dict[str, dict[str, Any]]:
    raw_cells: dict[str, dict[str, Any]] = {}
    max_row = min(sheet.max_row or MAX_RAW_ROWS, MAX_RAW_ROWS)
    max_column = min(sheet.max_column or MAX_RAW_COLUMNS, MAX_RAW_COLUMNS)

    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            cell = sheet.cell(row=row, column=column)
            raw_cells[cell.coordinate] = _cell_payload(cell)

    return raw_cells


def _row_values(
    sheet: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=False,
    ):
        rows.append([_json_value(cell.value) for cell in row])
    return rows


def _is_non_empty_row(values: list[Any]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in values)


def _detect_tables(sheet: Worksheet) -> dict[str, dict[str, Any]]:
    max_row = min(sheet.max_row or EMPTY_ROW_SCAN_LIMIT, EMPTY_ROW_SCAN_LIMIT)
    max_column = min(sheet.max_column or MAX_RAW_COLUMNS, MAX_RAW_COLUMNS)
    tables: dict[str, dict[str, Any]] = {}
    current_start: int | None = None
    current_rows: list[list[Any]] = []
    table_index = 1

    for row_index in range(1, max_row + 1):
        values = [
            _json_value(sheet.cell(row=row_index, column=column).value)
            for column in range(1, max_column + 1)
        ]
        if _is_non_empty_row(values):
            if current_start is None:
                current_start = row_index
                current_rows = []
            current_rows.append(values)
        elif current_start is not None:
            key = "left_work_table" if table_index == 1 else f"detected_table_{table_index}"
            tables[key] = {
                "range": f"A{current_start}:{get_column_letter(max_column)}{row_index - 1}",
                "data": current_rows,
            }
            table_index += 1
            current_start = None
            current_rows = []

    if current_start is not None:
        key = "left_work_table" if table_index == 1 else f"detected_table_{table_index}"
        tables[key] = {
            "range": f"A{current_start}:{get_column_letter(max_column)}{max_row}",
            "data": current_rows,
        }

    for table in sheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        tables[table.name] = {
            "range": table.ref,
            "data": _row_values(sheet, min_row, max_row, min_col, max_col),
        }

    if "detected_table_2" in tables and "metrics_table" not in tables:
        tables["metrics_table"] = tables["detected_table_2"]

    return tables


def _find_label_values(raw_cells: dict[str, dict[str, Any]], labels: list[str]) -> dict[str, Any]:
    found: dict[str, Any] = {}
    items = list(raw_cells.items())
    for coordinate, payload in items:
        value = payload["value"]
        if value is None:
            continue
        text = str(value).replace(" ", "").lower()
        for label in labels:
            normalized_label = label.replace(" ", "").lower()
            if normalized_label in text and label not in found:
                found[label] = {"cell": coordinate, "value": value}
    return found


def _extract_structured_data(
    sheet: Worksheet,
    raw_cells: dict[str, dict[str, Any]],
    tables: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    header_labels = ["품번", "품명", "모델", "라인", "작성", "개정", "공정", "제품"]
    metric_labels = [
        "tact",
        "tact time",
        "standard",
        "standard time",
        "lob",
        "eb",
        "capa",
        "생산량",
        "작업자",
    ]
    basic_labels = ["고객", "차종", "호기", "설비", "일자", "shift", "비고"]

    steps: list[dict[str, Any]] = []
    max_row = min(sheet.max_row or EMPTY_ROW_SCAN_LIMIT, EMPTY_ROW_SCAN_LIMIT)
    max_column = min(sheet.max_column or MAX_RAW_COLUMNS, MAX_RAW_COLUMNS)
    for row_index in range(1, max_row + 1):
        row_values = [
            _json_value(sheet.cell(row=row_index, column=column).value)
            for column in range(1, max_column + 1)
        ]
        if _is_non_empty_row(row_values):
            steps.append({"row": row_index, "values": row_values})

    return {
        "header": _find_label_values(raw_cells, header_labels),
        "steps": steps,
        "metrics": _find_label_values(raw_cells, metric_labels),
        "basic_info": _find_label_values(raw_cells, basic_labels),
        "detected_tables": list(tables.keys()),
    }


def _collect_warnings(
    sheet: Worksheet,
    raw_cells: dict[str, dict[str, Any]],
    merged_cells: list[str],
) -> list[str]:
    warnings: list[str] = []
    if merged_cells:
        warnings.append("병합 셀 발견")
    if any(payload["data_type"] == "f" for payload in raw_cells.values()):
        warnings.append("수식 발견")
    has_empty_cell = any(
        payload["value"] is None or str(payload["value"]).strip() == ""
        for payload in raw_cells.values()
    )
    if has_empty_cell:
        warnings.append("빈 셀 발견")

    max_row = min(sheet.max_row or EMPTY_ROW_SCAN_LIMIT, EMPTY_ROW_SCAN_LIMIT)
    max_column = min(sheet.max_column or MAX_RAW_COLUMNS, MAX_RAW_COLUMNS)
    for row_index in range(1, max_row + 1):
        values = [
            sheet.cell(row=row_index, column=column).value
            for column in range(1, max_column + 1)
        ]
        if not _is_non_empty_row(values):
            warnings.append(f"빈 행 발견: {row_index}")
            break

    return warnings


@router.post("/parse-excel/diagnose")
async def diagnose_standard_process_chart_excel(
    file: UploadFile = File(...),
) -> dict[str, Any]:
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel xlsx/xlsm/xltx/xltm 파일만 업로드할 수 있습니다.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="업로드된 파일이 비어 있습니다.",
        )

    try:
        workbook = load_workbook(io.BytesIO(contents), data_only=False, read_only=False)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Excel 파일을 파싱할 수 없습니다: {exc}",
        ) from exc

    sheet = workbook.active
    raw_cells = _extract_raw_cells(sheet)
    merged_cells = [str(cell_range) for cell_range in sheet.merged_cells.ranges]
    tables = _detect_tables(sheet)

    return {
        "filename": filename,
        "sheet_names": workbook.sheetnames,
        "active_sheet": sheet.title,
        "parsed_at": dt.datetime.now(dt.UTC).isoformat(),
        "dimensions": {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "scanned_raw_range": (
                f"A1:{get_column_letter(min(sheet.max_column or MAX_RAW_COLUMNS, MAX_RAW_COLUMNS))}"
                f"{min(sheet.max_row or MAX_RAW_ROWS, MAX_RAW_ROWS)}"
            ),
        },
        "raw_cells": raw_cells,
        "merged_cells": merged_cells,
        "tables": tables,
        "structured_data": _extract_structured_data(sheet, raw_cells, tables),
        "warnings": _collect_warnings(sheet, raw_cells, merged_cells),
    }
